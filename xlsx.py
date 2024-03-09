import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def write_to_cell(file_path, sheet_name, cell, text):
    # Create a new workbook if the file doesn't exist
    try:
        workbook = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        workbook = Workbook()
        workbook.save(file_path)

    # Check if the worksheet exists, if not create it
    if sheet_name not in workbook.sheetnames:
        worksheet = workbook.create_sheet(sheet_name)
    else:
        worksheet = workbook[sheet_name]

    # Write the text to the specified cell
    column, row = cell[0], cell[1]
    column_letter = get_column_letter(column)
    worksheet[f"{column_letter}{row}"] = text

    # Save the changes
    workbook.save(file_path)


def read_from_cell(file_path, sheet_name, cell):
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)

    # Select the worksheet
    worksheet = workbook[sheet_name]

    # Read the text from the specified cell
    column, row = cell[0], cell[1]
    column_letter = get_column_letter(column)
    text = worksheet[f"{column_letter}{row}"].value

    return text

write_to_cell("hola.xlsx", "Contactos", (1, 1), "Hola como estas ")

print(read_from_cell("hola.xlsx", "Contactos", (1, 1)))
